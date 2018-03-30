from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from random import *
from openpyxl import Workbook, load_workbook
from pathlib import Path
from .utils import wait_rand, wait_to_load, press_rand, login, load_linkedIn_links_from_file, driver

# options = webdriver.ChromeOptions()
# options.add_argument('headless')

# set the window size
# options.add_argument('window-size=1200x600')

# initialize the driver
# driver = webdriver.Chrome(chrome_options=options)

def log(account,status):
	xlx_file = Path("accounts_status_to_zoho.xlsx")
	#clean file before new import!?!?!?	

	if not xlx_file.is_file():	
		wb = Workbook()
		ws = wb.active
		ws = wb.create_sheet('accounts_to_zoho', 0)
		wb.save(xlx_file)
	else:
		wb = load_workbook(xlx_file)
		ws = wb['accounts_to_zoho']
		ws = wb.active

	data = [account['id'], status]
	ws.append(data)
	wb.save(filename = xlx_file)


def send_connect(account):	
	driver.get(account['link'])
	wait_to_load()
	
	try:
		available = driver.find_element_by_css_selector('div.profile-unavailable')	
		log(account,'account deleted')
		driver.get_screenshot_as_file('img/profile-unavailable-' + account['link'] + '.png')
		return
	except NoSuchElementException:
		print("Handled ERROR: ", NoSuchElementException)
		pass
	
	try:
		status = driver.find_element_by_css_selector('span.pv-s-profile-actions__label').text
		driver.get_screenshot_as_file('img/1_add-new-connect.png')
	except NoSuchElementException:
		print("Handled ERROR: ", NoSuchElementException)
		log(account,'account not found')
		driver.get_screenshot_as_file('img/profile-unavailable-' + account['link'] + '.png')
		return

	if status == "Connect":
		button = driver.find_element_by_css_selector('button.pv-s-profile-actions--connect')
		button.click()
		log(account,'Sent')
	elif status == "Message":
		log(account,'Connected')
		return
	elif status == "InMail":
		log(account,status)
		return
	elif status == "Pending":		
		log(account,status)
		return
	else:
		driver.get_screenshot_as_file('img/ERROR-' + account['link'] + '.png')
		return

	name = driver.find_element_by_css_selector('h1.pv-top-card-section__name').text
	text = "Hello " + name + ", can we connect on Linkedin?"

	button = driver.find_element_by_css_selector('button.mr1')
	button.click()

	wait_rand()

	textarea = driver.find_element_by_css_selector('textarea.send-invite__custom-message')
	textarea.send_keys(text)

	wait_rand()

	button = driver.find_element_by_css_selector('button.ml1')
	button.click()

	wait_to_load()

	driver.get_screenshot_as_file('img/2_' + name + '_sent.png')

	press_rand()


def connect():
	#not more than 40 for one session
	accounts = json.load(open('accounts.json'))

	for account in accounts:
		login(account['email'],account['password'])
		leads = load_linkedIn_links_from_file()	
		for lead in leads:
			try:
				send_connect(lead)
			except ValueError:
				print("Handled ERROR: ", ValueError)
				driver.get_screenshot_as_file('img/error-' + lead['link'] + '.png')
				continue	
	
	
	print("SUCCESS!!")



