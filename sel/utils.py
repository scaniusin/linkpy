from selenium import webdriver
from random import *
from openpyxl import load_workbook, Workbook
from pathlib import Path
from datetime import datetime
import json, math, time


options = webdriver.ChromeOptions()
options.add_argument('headless')

# set the window size
options.add_argument('window-size=1600x1200')

# initialize the driver
driver = webdriver.Chrome(chrome_options=options)

driver.set_page_load_timeout(20)
driver.implicitly_wait(10)
driver.set_script_timeout(10)

def wait_rand():
	time.sleep(uniform(2, 6))
def wait_to_load():
	time.sleep(5)
def wait_in_min(min):
	time.sleep(min * 60)
# try:
#    element_present = EC.presence_of_element_located((By.ID, 'whatever'))
#    WebDriverWait(driver, timeout).until(element_present)
# except TimeoutException:
#     print "Timed out waiting for page to load"

def log_file(action, variable):
	path = Path("log.txt")
	log_file = open(path,'a')
	current_time = '{0:%Y-%m-%d %H:%M:%S}'.format(datetime.now())
	if variable:
		text = str(current_time) + ' - ' + action + ' - ' + str(variable) + '\n'
	else:
		text = str(current_time) + ' - ' + action + ' - ' + '\n'
	log_file.write(text)
	print(text)
	log_file.close()



def press_rand():
	driver.refresh() #to change

	home = driver.find_element_by_css_selector('a[href="/feed/"]')
	myNetwork = driver.find_element_by_css_selector('a[href="/mynetwork/"]')
	jobs = driver.find_element_by_css_selector('a[href="/jobs/"]')
	messaging = driver.find_element_by_css_selector('a[href="/messaging/"]')
	notifications = driver.find_element_by_css_selector('a[href="/notifications/"]')

	wait_rand()

	buttons = [home, myNetwork, jobs, messaging, notifications]
	buttons[randrange(1, 5)].click()
	
	log_file("press rendom button", buttons)

	wait_rand()

	driver.get_screenshot_as_file('img/3-rand-press.png')


def login(email_input, password_input):
	# email_input = input('Insert account email: ')
	# password_input = input('Insert account password: ')
	
	driver.get("https://www.linkedin.com/")
	
	wait_to_load()

	email = driver.find_element_by_css_selector('#login-email')
	password = driver.find_element_by_css_selector('#login-password')
	login = driver.find_element_by_css_selector('#login-submit')

	log_file("login elements", str(email) + str(password) + str(login))

	email.send_keys(email_input)
	password.send_keys(password_input)
	
	login.click()

	log_file("login as: ", email_input)

	wait_to_load()

	driver.get_screenshot_as_file('img/0-login-success.png')

def logout():
	button = driver.find_element_by_css_selector('#nav-settings__dropdown-trigger')
	button.click()
	
	log_file("logout elements", button)

	wait_rand()

	button = driver.find_element_by_css_selector('a[href="/m/logout/"]')
	button.click()

	log_file("logout elements", button)

	wait_rand()

	driver.get("https://www.linkedin.com/")



def load_linkedIn_links_from_file(file, LIMIT, portion):
	wb = load_workbook(filename = file)
	ws = wb['Sheet 1']
	leads = []
	for row in ws.rows:
		# list_links.append(row[0].value)
		lead = {}
		lead['id'] = row[0].value
		lead['link'] = row[1].value
		leads.append(lead)
	if(math.ceil(len(leads)/LIMIT)-1 >= portion):
		chunks = [leads[x:x+LIMIT] for x in range(0, len(leads), LIMIT)]
		return chunks[portion]
	else:
		return []

def log(lead, status, account_email):
	xlx_file = Path("leads_status_for_import_zoho_" + account_email + ".xlsx")
	#clean file before new import!?!?!?	

	if not xlx_file.is_file():	
		wb = Workbook()
		ws = wb.active
		ws = wb.create_sheet('accounts_to_zoho', 0)
		wb.save(xlx_file)
	else:
		wb = load_workbook(xlx_file)
		ws = wb['accounts_to_zoho']
		ws = wb.active

	data = [lead['id'], status]
	ws.append(data)
	wb.save(filename = xlx_file)