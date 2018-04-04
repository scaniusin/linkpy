from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from random import *
from openpyxl import load_workbook, Workbook
from pathlib import Path
from time import gmtime, strftime
import json, math, time
from sel.utils import *

def write_to_file(url_id, url):
	xlx_out = Path("foridOUT.xlsx")
	
	wb = load_workbook(xlx_out)
	ws = wb['links']
	ws = wb.active

	data = [url_id, url]
	ws.append(data)
	wb.save(filename = xlx_out)

def getId():
	file = Path("forid.xlsx")
	wb = load_workbook(filename = file)

	ws = wb['Sheet1'] 
	leads = []
	for row in ws.rows:
		leads.append(row[0].value)

	count = 0
	for lead in leads:
		count+=1
		log_file("Number of lead #", count)
		if lead:
			try:
				driver.get(lead)
				log_file("load link", lead)
				wait_rand()
			except TimeoutException:
				write_to_file("not available", TimeoutException)
				continue
		else:
			write_to_file("not available", TimeoutException)
			continue
		


		try:
			available = driver.find_element_by_css_selector('div.profile-unavailable')	
			log_file("profile is unavailable", available)
			write_to_file("not available", lead)
			continue
		except NoSuchElementException:
			log_file("profile is available", NoSuchElementException)
			pass
		
		try:
			button = driver.find_element_by_css_selector('a[data-control-name="view_profile_in_sales_navigator"]')
			button.click()
			log_file("check if button sales available", button)
		except NoSuchElementException:
			write_to_file("not available", lead)
			log_file("no button element", NoSuchElementException)
			continue

		wait_to_load()

		driver.switch_to_window(driver.window_handles[1])

		url_id = driver.current_url

		log_file("result url", url_id)

		write_to_file(url_id, lead)

		driver.close()
		driver.switch_to_window(driver.window_handles[0])

		wait_rand()
	
		if count%50 == 0:
			log_file("sleep 2 min", count)
			time.sleep(120)

try:
	log_file("\n\n START of Script get ID \n", None)

	login("aksennyaa@gmail.com", "openmenow15")

	getId()

	log_file("SUCCESS", None)
except Exception as ex:
    log_file("ERROR", format(ex))
    driver.get_screenshot_as_file('img/ERROR.png')
    print('Error: {}'.format(ex))

	






